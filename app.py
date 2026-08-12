import io
import os
import re
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components

# Page Configuration
st.set_page_config(page_title="FTS Management Portal", layout="centered")

# Direct Shareable Download URL for your SharePoint File
SHAREPOINT_DIRECT_DOWNLOAD_URL = "https://tilaknagarindustries-my.sharepoint.com/:x:/g/personal/andebnath_tilind_com/IQBXwlCv60OJSoXvUAk-7j7FATrDyVAIk_UoFRX4p1k2pWw?download=1"

# Local search fallback paths (for offline desktop use)
PRIMARY_XLSX = r"C:\Users\admin\OneDrive - Tilak Nagar Industries Ltd\Documents\App Data\FTS Calculation Update.xlsx"
LOCAL_SEARCH_PATHS = [
    PRIMARY_XLSX,
    "FTS Calculation Update.xlsx",
    "FTS Calculator.xlsx",
]

BRANDS = ["IBDC", "MHW", "BLGLM", "BLGOR", "MHFB", "SMG", "SMGP", "SIW", "Monarch"]

POINTS_CONFIG = {
    "IBDC": (20, 5),
    "MHW": (40, 10),
    "BLGLM": (10, 5),
    "BLGOR": (10, 5),
    "MHFB": (30, 10),
    "SMG": (150, 75),
    "SMGP": (150, 75),
    "SIW": (300, 150),
    "Monarch": (300, 150),
}


@st.cache_data(ttl=10)
def fetch_sharepoint_excel_bytes():
    """Downloads live Excel workbook bytes directly from SharePoint share link."""
    try:
        res = requests.get(SHAREPOINT_DIRECT_DOWNLOAD_URL, timeout=10)
        if res.status_code == 200:
            return res.content
    except Exception:
        pass

    # Fallback to local desktop file if offline
    for path in LOCAL_SEARCH_PATHS:
        if os.path.exists(path):
            with open(path, "rb") as f:
                return f.read()

    return None


@st.cache_data(ttl=10)
def load_outlet_master():
    """Reads Outlet Master sheet directly from live SharePoint workbook."""
    excel_bytes = fetch_sharepoint_excel_bytes()
    if excel_bytes:
        try:
            return pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Outlet Master")
        except Exception:
            pass

    return pd.DataFrame(
        {
            "Lic ID": ["LIC1001", "LIC1002", "LIC1003", "LIC1004"],
            "Outlet Name": ["Royal Wine Shop", "Grand Spirits", "City Bar", "Metro Wines"],
            "Group/Individual": ["Group", "Individual", "Group", "Individual"],
            "ASM": ["ASM - North", "ASM - North", "ASM - South", "ASM - South"],
            "TSE": ["TSE - Kolkata 1", "TSE - Kolkata 2", "TSE - Howrah", "TSE - Howrah"],
        }
    )


@st.cache_data(ttl=5)
def load_full_enrollment_raw():
    """Reads raw Enrollment sheet directly from live SharePoint workbook."""
    excel_bytes = fetch_sharepoint_excel_bytes()
    if excel_bytes:
        try:
            xl = pd.ExcelFile(io.BytesIO(excel_bytes))
            sheet_name = "Enrollment" if "Enrollment" in xl.sheet_names else "Enrolment"
            return pd.read_excel(io.BytesIO(excel_bytes), sheet_name=sheet_name, header=None)
        except Exception:
            pass
    return pd.DataFrame()


def update_row_references(formula_str, old_row, new_row):
    """Adjusts cell row numbers in Excel formulas."""
    if not isinstance(formula_str, str) or not formula_str.startswith("="):
        return formula_str

    def replace_row(match):
        col = match.group(1)
        row = int(match.group(2))
        if row == old_row:
            return f"{col}{new_row}"
        return match.group(0)

    pattern = r"([A-Za-z]+)(\d+)"
    return re.sub(pattern, replace_row, formula_str)


def append_enrolment_to_bytes(row_values):
    """Appends enrolment record and drags formulas down in-memory."""
    excel_bytes = fetch_sharepoint_excel_bytes()
    if not excel_bytes:
        return False, "Could not load base Excel file.", None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        sheet_name = "Enrollment" if "Enrollment" in wb.sheetnames else "Enrolment"
        sheet = wb[sheet_name]

        new_row_idx = sheet.max_row + 1
        if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
            new_row_idx = 1

        for col_idx, val in enumerate(row_values, start=1):
            sheet.cell(row=new_row_idx, column=col_idx, value=val)

        prev_row_idx = new_row_idx - 1
        if prev_row_idx >= 3:
            for col_idx in range(len(row_values) + 1, sheet.max_column + 1):
                prev_cell = sheet.cell(row=prev_row_idx, column=col_idx)
                prev_val = prev_cell.value
                if isinstance(prev_val, str) and prev_val.startswith("="):
                    new_formula = update_row_references(prev_val, prev_row_idx, new_row_idx)
                    sheet.cell(row=new_row_idx, column=col_idx, value=new_formula)

        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        return True, "Successfully registered enrolment entry!", out_buffer.getvalue()
    except Exception as e:
        return False, f"Error processing file: {e}", None


def get_calculated_tour(total_points):
    if total_points >= 12000:
        return "4N/ 5D Mauritius"
    elif total_points >= 9000:
        return "6N/ 7D Ladakh"
    elif total_points >= 6000:
        return "4N/ 5D Manali"
    else:
        return "Not Eligible"


# ==========================================
# SIDEBAR NAVIGATION & EXCEL DOWNLOAD
# ==========================================
st.sidebar.title("📌 FTS Portal Menu")

page = st.sidebar.radio(
    "Select Page",
    ["Ach as per enrolment", "FTS Calculator", "Enrol party for FTS"],
)

st.sidebar.markdown("---")
st.sidebar.subheader("💾 Master Excel File")

master_bytes = fetch_sharepoint_excel_bytes()
if master_bytes:
    st.sidebar.download_button(
        label="📥 Download Master Excel Workbook",
        data=master_bytes,
        file_name="FTS Calculation Update.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ==========================================
# PAGE 1: ACH AS PER ENROLMENT
# ==========================================
if page == "Ach as per enrolment":
    st.title("📊 Achievement vs Enrolment Report")
    st.write("Live achievement tracking against committed party targets:")

    df_raw = load_full_enrollment_raw()

    if df_raw.empty or len(df_raw) < 3:
        st.warning("No enrolment data available.")
    else:
        main_cols = [
            "Asm", "TSE", "Lic ID/Group", "Outlet Name", "Group/Individual",
            "4N/ 5D Mauritius", "6N/ 7D Ladakh", "4N/ 5D Manali",
            "Total Ticket", "Total Point Required", "Ach Point", "Balance"
        ]

        df_records = df_raw.iloc[3:, :12].copy()
        df_records.columns = main_cols
        df_records.dropna(how="all", inplace=True)

        f_col1, f_col2, f_col3 = st.columns(3)

        asm_opts = ["All"] + sorted(df_records["Asm"].dropna().astype(str).unique().tolist())
        selected_asm = f_col1.selectbox("ASM Filter", asm_opts)

        filtered_df = df_records.copy()
        raw_filtered = df_raw.iloc[3:].copy()

        if selected_asm != "All":
            mask = filtered_df["Asm"].astype(str) == selected_asm
            filtered_df = filtered_df[mask]
            raw_filtered = raw_filtered[mask]

        tse_opts = ["All"] + sorted(filtered_df["TSE"].dropna().astype(str).unique().tolist())
        selected_tse = f_col2.selectbox("TSE Filter", tse_opts)

        if selected_tse != "All":
            mask = filtered_df["TSE"].astype(str) == selected_tse
            filtered_df = filtered_df[mask]
            raw_filtered = raw_filtered[mask]

        outlet_opts = ["All"] + sorted(filtered_df["Outlet Name"].dropna().astype(str).unique().tolist())
        selected_outlet = f_col3.selectbox("Outlet Name Filter", outlet_opts)

        if selected_outlet != "All":
            mask = filtered_df["Outlet Name"].astype(str) == selected_outlet
            filtered_df = filtered_df[mask]
            raw_filtered = raw_filtered[mask]

        st.markdown("---")

        if selected_outlet != "All" and len(filtered_df) == 1:
            row_idx = filtered_df.index[0]
            raw_row = df_raw.loc[row_idx]

            st.markdown(f"### 📍 Enrollment Summary — {selected_outlet}")
            st.dataframe(filtered_df, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### 📈 Monthly Achievement Breakdown")

            def build_month_df(start_col, title):
                vals = raw_row.iloc[start_col : start_col + 9].values
                vals = [0 if pd.isna(v) else v for v in vals]
                return pd.DataFrame([vals], columns=BRANDS)

            st.markdown("##### **Aug'26 Secondary**")
            st.dataframe(build_month_df(12, "Aug'26 Secondary"), use_container_width=True, hide_index=True)

            st.markdown("##### **Sep'26 Secondary**")
            st.dataframe(build_month_df(21, "Sep'26 Secondary"), use_container_width=True, hide_index=True)

            st.markdown("##### **Oct'26 Secondary**")
            st.dataframe(build_month_df(30, "Oct'26 Secondary"), use_container_width=True, hide_index=True)

            st.markdown("##### **Aug'26 Tertiary**")
            st.dataframe(build_month_df(40, "Aug'26 Tertiary"), use_container_width=True, hide_index=True)

            st.markdown("##### **Sep'26 Tertiary**")
            st.dataframe(build_month_df(49, "Sep'26 Tertiary"), use_container_width=True, hide_index=True)

            st.markdown("##### **Oct'26 Tertiary**")
            st.dataframe(build_month_df(58, "Oct'26 Tertiary"), use_container_width=True, hide_index=True)

        else:
            st.markdown("### 📋 Enrollment Overview")
            st.dataframe(filtered_df, use_container_width=True, hide_index=True)

# ==========================================
# PAGE 2: FTS CALCULATOR
# ==========================================
elif page == "FTS Calculator":
    st.title("FTS Calculator")

    outlet_name_calc = st.text_input("Outlet Name", value="", placeholder="Enter Outlet Name here...")
    st.subheader("Data Input — Aug'26 to Oct'26 Plan")

    inputs = {}
    total_calculated_points = 0
    total_sec = 0
    total_tert = 0

    for b in BRANDS:
        c1, c2 = st.columns(2)
        sec_val = c1.number_input(f"{b} (Secondary)", min_value=0, value=0, step=1, key=f"{b}_sec")
        tert_val = c2.number_input(f"{b} (Tertiary)", min_value=0, value=0, step=1, key=f"{b}_tert")

        inputs[b] = (sec_val, tert_val)
        total_sec += sec_val
        total_tert += tert_val

        pts_sec, pts_tert = POINTS_CONFIG[b]
        total_calculated_points += (sec_val * pts_sec) + (tert_val * pts_tert)

    calculated_tour = get_calculated_tour(total_calculated_points)
    outlet_display_str = outlet_name_calc.strip() if outlet_name_calc.strip() else "N/A"

    st.markdown("---")
    st.subheader("Summary Report")

    rows_html = ""
    for b in BRANDS:
        sec, tert = inputs[b]
        sec_disp = str(sec) if sec > 0 else ""
        tert_disp = str(tert) if tert > 0 else ""
        rows_html += f"""
        <tr>
            <td class="lbl-brand">{b}</td>
            <td class="cell-val">{sec_disp}</td>
            <td class="cell-val">{tert_disp}</td>
        </tr>
        """

    full_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
    <style>
        * {{ box-sizing: border-box; }}
        body {{ margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; background-color: transparent; display: flex; flex-direction: column; align-items: center; }}
        #reportCaptureArea {{ padding: 10px; background-color: white; width: 100%; max-width: 400px; border-radius: 6px; }}
        .fts-table {{ width: 100%; margin: 0 auto; border-collapse: collapse; font-weight: bold; font-size: clamp(11px, 3.2vw, 14px); text-align: center; border: 2px solid #000; cursor: pointer; user-select: none; }}
        .fts-table th, .fts-table td {{ border: 1px solid #000; padding: 6px 4px; word-break: break-word; vertical-align: middle; }}
        .hdr-outlet {{ background-color: #FFC000; color: black; text-align: center; font-size: clamp(12px, 3.5vw, 15px); }}
        .hdr-main {{ background-color: #002060; color: white; text-align: center; }}
        .hdr-month {{ background-color: #C6EFCE; color: #006100; text-align: center; }}
        .lbl-brand {{ background-color: #7030A0; color: white; text-align: left; padding-left: 10px; width: 40%; }}
        .cell-val {{ background-color: white; width: 30%; text-align: center; }}
        .row-total {{ background-color: #00B0F0; color: black; text-align: center; }}
        .row-points {{ background-color: #E2EFDA; color: black; text-align: center; }}
        .row-tour {{ background-color: #003300; color: white; text-align: center; }}
        .btn-share {{ background-color: #25D366; color: white; border: none; padding: 12px 20px; font-size: 14px; font-weight: bold; border-radius: 6px; cursor: pointer; width: 100%; max-width: 320px; margin-top: 10px; display: flex; align-items: center; justify-content: center; gap: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.2); }}
    </style>
    </head>
    <body>
    <div id="reportCaptureArea">
        <table class="fts-table" id="summaryTable">
            <thead>
                <tr class="hdr-outlet"><td colspan="3">Outlet Name: {outlet_display_str}</td></tr>
                <tr class="hdr-main"><th>Brand</th><th>Secondary</th><th>Tertiary</th></tr>
                <tr class="hdr-month"><td colspan="3">Aug'26 to Oct'26 Plan</td></tr>
            </thead>
            <tbody>
                {rows_html}
                <tr class="row-total"><td>Total</td><td>{total_sec}</td><td>{total_tert}</td></tr>
                <tr class="row-points"><td>Total Point</td><td colspan="2">{total_calculated_points}</td></tr>
                <tr class="row-tour"><td>Calculated Tour</td><td colspan="2">{calculated_tour}</td></tr>
            </tbody>
        </table>
    </div>
    <button class="btn-share" onclick="shareReportImage()">📲 Share Report Image</button>

    <script>
    async function shareReportImage() {{
        const area = document.getElementById('reportCaptureArea');
        try {{
            const canvas = await html2canvas(area, {{ scale: 2, useCORS: true }});
            canvas.toBlob(async function(blob) {{
                const fileName = 'FTS_Report_{outlet_display_str.replace(" ", "_")}.png';
                const file = new File([blob], fileName, {{ type: 'image/png' }});
                if (navigator.canShare && navigator.canShare({{ files: [file] }})) {{
                    await navigator.share({{
                        files: [file],
                        title: 'FTS Summary Report — {outlet_display_str}',
                        text: 'FTS Summary Report image for {outlet_display_str}'
                    }});
                }} else {{
                    const link = document.createElement('a');
                    link.download = fileName;
                    link.href = canvas.toDataURL('image/png');
                    link.click();
                }}
            }}, 'image/png');
        }} catch (err) {{
            console.error("Error generating image report:", err);
        }}
    }}
    </script>
    </body>
    </html>
    """

    components.html(full_html, height=750, scrolling=False)

    def generate_snapshot_image():
        fig, ax = plt.subplots(figsize=(4, 6.5), dpi=200)
        ax.axis("off")

        table_data = [
            [f"Outlet Name: {outlet_display_str}", "", ""],
            ["Brand", "Secondary", "Tertiary"],
            ["Aug'26 to Oct'26 Plan", "", ""],
        ]
        cell_colors = [
            ["#FFC000", "#FFC000", "#FFC000"],
            ["#002060", "#002060", "#002060"],
            ["#C6EFCE", "#C6EFCE", "#C6EFCE"],
        ]

        for b in BRANDS:
            sec, tert = inputs[b]
            table_data.append([b, str(sec) if sec > 0 else "", str(tert) if tert > 0 else ""])
            cell_colors.append(["#7030A0", "#FFFFFF", "#FFFFFF"])

        table_data.append(["Total", str(total_sec), str(total_tert)])
        cell_colors.append(["#00B0F0", "#00B0F0", "#00B0F0"])

        table_data.append(["Total Point", str(total_calculated_points), ""])
        cell_colors.append(["#E2EFDA", "#E2EFDA", "#E2EFDA"])

        table_data.append(["Calculated Tour", str(calculated_tour), ""])
        cell_colors.append(["#003300", "#003300", "#003300"])

        tab = ax.table(cellText=table_data, cellColours=cell_colors, loc="center", cellLoc="center")
        tab.scale(1, 1.3)

        buf = io.BytesIO()
        plt.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.1)
        buf.seek(0)
        plt.close(fig)
        return buf

    st.markdown("---")
    img_buf = generate_snapshot_image()
    st.download_button(
        label="📷 Download Table Snap (PNG)",
        data=img_buf,
        file_name=f"FTS_Calculator_{outlet_display_str.replace(' ', '_')}.png",
        mime="image/png",
    )

# ==========================================
# PAGE 3: ENROL PARTY FOR FTS
# ==========================================
elif page == "Enrol party for FTS":
    st.title("📝 Enrol Party for FTS")
    st.write("Fill in party enrolment details below:")

    df_master = load_outlet_master()

    lic_col = df_master.columns[1] if len(df_master.columns) > 1 else df_master.columns[0]
    outlet_col = df_master.columns[5] if len(df_master.columns) > 5 else df_master.columns[0]
    group_ind_col = df_master.columns[6] if len(df_master.columns) > 6 else df_master.columns[0]
    asm_col = df_master.columns[8] if len(df_master.columns) > 8 else df_master.columns[0]
    tse_col = df_master.columns[15] if len(df_master.columns) > 15 else df_master.columns[0]

    asm_list = ["All"] + sorted(df_master[asm_col].dropna().astype(str).unique().tolist())
    selected_asm = st.selectbox("Select ASM (Col I)", asm_list)

    filtered_df = df_master.copy()
    if selected_asm != "All":
        filtered_df = filtered_df[filtered_df[asm_col].astype(str) == selected_asm]

    tse_list = ["-- Select TSE --"] + sorted(filtered_df[tse_col].dropna().astype(str).unique().tolist())
    selected_tse = st.selectbox("Select TSE (Col P) *", tse_list)

    if selected_tse != "-- Select TSE --":
        filtered_df = filtered_df[filtered_df[tse_col].astype(str) == selected_tse]

    group_ind_list = ["All"] + sorted(filtered_df[group_ind_col].dropna().astype(str).unique().tolist())
    selected_group_ind = st.selectbox("Select Group / Individual (Col G)", group_ind_list)

    if selected_group_ind != "All":
        filtered_df = filtered_df[filtered_df[group_ind_col].astype(str) == selected_group_ind]

    outlet_options = filtered_df[outlet_col].dropna().astype(str).unique().tolist()

    with st.form("party_enrolment_form", clear_on_submit=True):
        selected_outlet = st.selectbox(
            "Party / Outlet Name * (Col F)",
            options=outlet_options if outlet_options else ["No Outlet Found"],
        )

        auto_lic_id = ""
        auto_group_type = ""
        if selected_outlet and selected_outlet != "No Outlet Found":
            matched_row = filtered_df[filtered_df[outlet_col].astype(str) == selected_outlet]
            if not matched_row.empty:
                auto_lic_id = str(matched_row.iloc[0][lic_col])
                auto_group_type = str(matched_row.iloc[0][group_ind_col])

        c1, c2 = st.columns(2)
        with c1:
            st.text_input("Lic ID / Group (Auto-fetched from Col B)", value=auto_lic_id, disabled=True)
        with c2:
            st.text_input("Group / Individual (Auto-fetched from Col G)", value=auto_group_type, disabled=True)

        st.markdown("---")
        st.markdown("**Enter Tour Target Quantities (Tickets)**")

        q_col1, q_col2, q_col3 = st.columns(3)

        with q_col1:
            st.markdown("##### 12,000 Pts")
            qty_mauritius = st.number_input("4N/ 5D Mauritius (Qty)", min_value=0, value=0, step=1, key="qty_mauritius")

        with q_col2:
            st.markdown("##### 9,000 Pts")
            qty_ladakh = st.number_input("6N/ 7D Ladakh (Qty)", min_value=0, value=0, step=1, key="qty_ladakh")

        with q_col3:
            st.markdown("##### 6,000 Pts")
            qty_manali = st.number_input("4N/ 5D Manali (Qty)", min_value=0, value=0, step=1, key="qty_manali")

        submit_btn = st.form_submit_button("🚀 Submit Enrolment")

    if submit_btn:
        if selected_tse == "-- Select TSE --":
            st.error("❌ TSE Selection is mandatory! Please select a specific TSE.")
        elif not selected_outlet or selected_outlet == "No Outlet Found":
            st.error("Please select a valid Outlet.")
        elif qty_mauritius == 0 and qty_ladakh == 0 and qty_manali == 0:
            st.error("Please enter quantity for at least one Tour Target option.")
        else:
            total_tickets = qty_mauritius + qty_ladakh + qty_manali
            total_point_required = (qty_mauritius * 12000) + (qty_ladakh * 9000) + (qty_manali * 6000)

            row_data = [
                selected_asm if selected_asm != "All" else "",
                selected_tse,
                auto_lic_id,
                selected_outlet,
                auto_group_type,
                qty_mauritius,
                qty_ladakh,
                qty_manali,
                total_tickets,
                total_point_required,
            ]

            success, msg, updated_bytes = append_enrolment_to_bytes(row_data)
            if success and updated_bytes:
                st.success("✅ Enrolment recorded successfully!")
                st.download_button(
                    label="📥 Download Updated Excel File",
                    data=updated_bytes,
                    file_name="FTS Calculation Update.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.error(f"❌ {msg}")
